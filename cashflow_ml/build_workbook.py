"""
Build a formatted Excel workbook from the bank statement CSV.

Everything that can be a formula is a formula: change a figure on the
Transactions sheet and the P&L, the monthly cash flow and the dashboard all
follow. That is the difference between a spreadsheet and a printout.

Run:
    python build_workbook.py
    python build_workbook.py path/to/bank.csv
"""

from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.table import Table, TableStyleInfo

BASE_DIR = Path(__file__).resolve().parent
DEFAULT_CSV = BASE_DIR / "data" / "bank_company.csv"
OUTPUT = BASE_DIR / "output" / "cashflow_workbook.xlsx"

FONT = "Arial"
INK = "1A1D21"
MUTED = "6B7280"
RULE = "BFC4CB"
BAND = "F2F4F6"
HEAD = "1F3B57"

# Money in black, negatives in brackets, a zero shown as a dash.
GBP = '£#,##0.00;(£#,##0.00);"-"'
GBP0 = '£#,##0;(£#,##0);"-"'
PCT = '0.0%'
DATE = 'dd/mm/yyyy'

# Keyword -> category. Held on its own sheet so it can be edited without
# touching a formula; the Transactions sheet looks the category up from here.
CATEGORY_RULES: list[tuple[str, str, str]] = [
    ("RENT", "Rental income", "Income"),
    ("MORTGAGE", "Mortgage interest", "Expense"),
    ("AGENT", "Managing agent fees", "Expense"),
    ("INSURANCE", "Insurance", "Expense"),
    ("ACCOUNTANCY", "Professional fees", "Expense"),
    ("REMUNERATION", "Directors remuneration", "Expense"),
    ("REPAIR", "Repairs and maintenance", "Expense"),
]


def read_csv(path: Path) -> list[dict[str, str]]:
    import csv

    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def number(value: str) -> float:
    return float(value) if value.strip() else 0.0


# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------
def title_block(sheet, title: str, subtitle: str, width: int) -> None:
    sheet["A1"] = title
    sheet["A1"].font = Font(name=FONT, size=15, bold=True, color=INK)
    sheet["A2"] = subtitle
    sheet["A2"].font = Font(name=FONT, size=9, color=MUTED)
    sheet.row_dimensions[1].height = 22
    sheet.row_dimensions[3].height = 6


def header_row(sheet, row: int, headings: list[str]) -> None:
    for column, text in enumerate(headings, start=1):
        cell = sheet.cell(row=row, column=column, value=text)
        cell.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEAD)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
    sheet.row_dimensions[row].height = 26


def widths(sheet, sizes: dict[str, int]) -> None:
    for letter, size in sizes.items():
        sheet.column_dimensions[letter].width = size


def printable(sheet, header: int | None = 4) -> None:
    """Landscape, squeezed onto one page wide, header repeated on each page.

    Without this a wide sheet prints its last columns on a second page, which
    is how a tidy workbook turns into a mess on someone's desk.
    """
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    sheet.print_options.horizontalCentered = False
    if header:
        sheet.print_title_rows = f"{header}:{header}"


def thin_rule(sheet, row: int, columns: int, weight: str = "thin") -> None:
    side = Side(style=weight, color=RULE)
    for column in range(1, columns + 1):
        cell = sheet.cell(row=row, column=column)
        cell.border = Border(top=side)


# ---------------------------------------------------------------------------
# Sheets
# ---------------------------------------------------------------------------
def build_categories(sheet) -> None:
    title_block(sheet, "Category rules",
                "Edit these to change how transactions are categorised. "
                "The first keyword found in the description wins.", 3)
    header_row(sheet, 4, ["Keyword (upper case)", "Category", "Type"])

    for index, (keyword, category, kind) in enumerate(CATEGORY_RULES):
        row = 5 + index
        sheet.cell(row=row, column=1, value=keyword).font = Font(
            name=FONT, size=10, color="0000FF")     # blue = you may edit
        sheet.cell(row=row, column=2, value=category).font = Font(
            name=FONT, size=10, color="0000FF")
        sheet.cell(row=row, column=3, value=kind).font = Font(
            name=FONT, size=10, color="0000FF")
        for column in range(1, 4):
            sheet.cell(row=row, column=column).fill = PatternFill(
                "solid", fgColor="FFFFCC")

    note = 5 + len(CATEGORY_RULES) + 1
    sheet.cell(row=note, column=1,
               value="Blue text on yellow = input cells. Everything else in "
                     "this workbook is calculated.")
    sheet.cell(row=note, column=1).font = Font(name=FONT, size=9, italic=True,
                                               color=MUTED)
    widths(sheet, {"A": 24, "B": 26, "C": 12})


def build_transactions(sheet, rows: list[dict[str, str]]) -> int:
    title_block(sheet, "Transactions",
                "One row per bank line. Columns H to K are calculated - "
                "do not type over them.", 11)
    header_row(sheet, 4, [
        "Date", "Description", "Counterparty", "Property", "Document",
        "Money in", "Money out", "Amount", "Direction", "Category", "Month",
    ])

    first = 5
    last = first + len(rows) - 1
    rule_end = 4 + len(CATEGORY_RULES)

    for index, record in enumerate(rows):
        row = first + index
        date = datetime.strptime(record["date"].strip(), "%d/%m/%Y")

        sheet.cell(row=row, column=1, value=date).number_format = DATE
        sheet.cell(row=row, column=2, value=record["description"].strip())
        sheet.cell(row=row, column=3, value=record["counterparty"].strip())
        sheet.cell(row=row, column=4,
                   value=record["property"].strip() or "Company-wide")
        sheet.cell(row=row, column=5, value=record["document"].strip())
        sheet.cell(row=row, column=6, value=number(record["paid in"]))
        sheet.cell(row=row, column=7, value=number(record["paid out"]))

        # Signed amount: money in positive, money out negative.
        sheet.cell(row=row, column=8, value=f"=F{row}-G{row}")
        sheet.cell(row=row, column=9, value=f'=IF(H{row}>=0,"Cash in","Cash out")')

        # Category lookup. SEARCH finds the keyword anywhere in the
        # description; MATCH finds the first rule that hits. The INDEX(...,0)
        # wrapper makes it work without Ctrl+Shift+Enter.
        sheet.cell(row=row, column=10, value=(
            f'=IFERROR(INDEX(Categories!$B$5:$B${rule_end},'
            f'MATCH(1,INDEX(--ISNUMBER(SEARCH(Categories!$A$5:$A${rule_end},'
            f'UPPER($B{row}))),0),0)),"Uncategorised")'
        ))
        sheet.cell(row=row, column=11, value=f'=TEXT(A{row},"yyyy-mm")')

        for column in range(1, 12):
            cell = sheet.cell(row=row, column=column)
            cell.font = Font(name=FONT, size=10, color=INK)
            if index % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=BAND)
        for column in (6, 7, 8):
            sheet.cell(row=row, column=column).number_format = GBP

    total = last + 1
    sheet.cell(row=total, column=5, value="Total").font = Font(
        name=FONT, size=10, bold=True)
    for column in (6, 7, 8):
        letter = get_column_letter(column)
        cell = sheet.cell(row=total, column=column,
                          value=f"=SUM({letter}{first}:{letter}{last})")
        cell.font = Font(name=FONT, size=10, bold=True)
        cell.number_format = GBP
    thin_rule(sheet, total, 11, "medium")

    widths(sheet, {"A": 12, "B": 38, "C": 30, "D": 15, "E": 22, "F": 13,
                   "G": 13, "H": 13, "I": 11, "J": 22, "K": 10})
    sheet.freeze_panes = "A5"

    # A real Excel table: gives filter buttons and banded rows.
    # Table name deliberately differs from the sheet name, or
    # "Transactions!" in a formula is ambiguous.
    table = Table(displayName="tblTransactions", ref=f"A4:K{last}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1", showRowStripes=False)
    sheet.add_table(table)
    return last


def build_monthly(sheet, months: list[str], last: int) -> None:
    title_block(sheet, "Monthly cash flow",
                "Totalled straight from the Transactions sheet with SUMIFS.", 5)
    header_row(sheet, 4, ["Month", "Cash in", "Cash out", "Net movement",
                          "Running total"])

    src = f"Transactions!$K$5:$K${last}"
    money_in = f"Transactions!$F$5:$F${last}"
    money_out = f"Transactions!$G$5:$G${last}"

    for index, month in enumerate(months):
        row = 5 + index
        sheet.cell(row=row, column=1, value=month)
        sheet.cell(row=row, column=2,
                   value=f'=SUMIFS({money_in},{src},$A{row})')
        sheet.cell(row=row, column=3,
                   value=f'=SUMIFS({money_out},{src},$A{row})')
        sheet.cell(row=row, column=4, value=f"=B{row}-C{row}")
        sheet.cell(row=row, column=5,
                   value=f"=D{row}" if index == 0 else f"=E{row - 1}+D{row}")
        for column in range(1, 6):
            cell = sheet.cell(row=row, column=column)
            cell.font = Font(name=FONT, size=10, color=INK)
            if column > 1:
                cell.number_format = GBP
            if index % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=BAND)

    total = 5 + len(months)
    sheet.cell(row=total, column=1, value="Total").font = Font(
        name=FONT, size=10, bold=True)
    for column in (2, 3, 4):
        letter = get_column_letter(column)
        cell = sheet.cell(row=total, column=column,
                          value=f"=SUM({letter}5:{letter}{total - 1})")
        cell.font = Font(name=FONT, size=10, bold=True)
        cell.number_format = GBP
    thin_rule(sheet, total, 5, "medium")
    widths(sheet, {"A": 12, "B": 15, "C": 15, "D": 15, "E": 16})
    sheet.freeze_panes = "A5"


def build_profit_and_loss(sheet, last: int) -> None:
    title_block(sheet, "Profit and loss",
                "Cash basis: receipts and payments as they cleared the bank. "
                "Not a statutory profit - no accruals, prepayments, "
                "depreciation or tax.", 3)
    header_row(sheet, 4, ["", "Transactions", "Amount"])

    amounts = f"Transactions!$H$5:$H${last}"
    categories = f"Transactions!$J$5:$J${last}"
    income = [r for r in CATEGORY_RULES if r[2] == "Income"]
    expense = [r for r in CATEGORY_RULES if r[2] == "Expense"]

    row = 5
    sheet.cell(row=row, column=1, value="INCOME").font = Font(
        name=FONT, size=9, bold=True, color=MUTED)
    row += 1
    income_first = row
    for _, category, _kind in income:
        sheet.cell(row=row, column=1, value=f"    {category}")
        sheet.cell(row=row, column=2,
                   value=f'=COUNTIFS({categories},"{category}")')
        sheet.cell(row=row, column=3,
                   value=f'=SUMIFS({amounts},{categories},"{category}")')
        row += 1
    income_total = row
    sheet.cell(row=income_total, column=1, value="Total income")
    sheet.cell(row=income_total, column=3,
               value=f"=SUM(C{income_first}:C{income_total - 1})")

    row += 2
    sheet.cell(row=row, column=1, value="EXPENDITURE").font = Font(
        name=FONT, size=9, bold=True, color=MUTED)
    row += 1
    expense_first = row
    for _, category, _kind in expense:
        sheet.cell(row=row, column=1, value=f"    {category}")
        sheet.cell(row=row, column=2,
                   value=f'=COUNTIFS({categories},"{category}")')
        sheet.cell(row=row, column=3,
                   value=f'=SUMIFS({amounts},{categories},"{category}")')
        row += 1
    expense_total = row
    sheet.cell(row=expense_total, column=1, value="Total expenditure")
    sheet.cell(row=expense_total, column=3,
               value=f"=SUM(C{expense_first}:C{expense_total - 1})")

    result = expense_total + 2
    sheet.cell(row=result, column=1, value="NET RESULT FOR THE PERIOD")
    sheet.cell(row=result, column=3,
               value=f"=C{income_total}+C{expense_total}")

    for r in range(5, result + 1):
        for column in range(1, 4):
            cell = sheet.cell(row=r, column=column)
            if cell.font.color is None or cell.font.color.rgb != "FF6B7280":
                cell.font = Font(name=FONT, size=10, color=INK)
        sheet.cell(row=r, column=3).number_format = GBP

    for r in (income_total, expense_total):
        for column in range(1, 4):
            sheet.cell(row=r, column=column).font = Font(
                name=FONT, size=10, bold=True, color=INK)
            sheet.cell(row=r, column=column).fill = PatternFill(
                "solid", fgColor=BAND)
        thin_rule(sheet, r, 3)

    for column in range(1, 4):
        sheet.cell(row=result, column=column).font = Font(
            name=FONT, size=11, bold=True, color=INK)
    thin_rule(sheet, result, 3, "medium")
    thin_rule(sheet, result + 1, 3, "medium")

    note = result + 3
    sheet.cell(row=note, column=1, value=(
        "Assumption: every bank line is revenue in nature. No capital items, "
        "loan drawdowns or repayments appear in this statement."))
    sheet.cell(row=note, column=1).font = Font(name=FONT, size=9, italic=True,
                                               color=MUTED)
    widths(sheet, {"A": 34, "B": 14, "C": 16})


def build_by_company(sheet, companies: list[str], last: int) -> None:
    title_block(sheet, "Accumulated by company",
                "Everything paid to or received from each counterparty.", 6)
    header_row(sheet, 4, ["Company", "Transactions", "Money in", "Money out",
                          "Accumulated", "Share of direction"])

    names = f"Transactions!$C$5:$C${last}"
    money_in = f"Transactions!$F$5:$F${last}"
    money_out = f"Transactions!$G$5:$G${last}"
    amounts = f"Transactions!$H$5:$H${last}"

    for index, company in enumerate(companies):
        row = 5 + index
        sheet.cell(row=row, column=1, value=company)
        sheet.cell(row=row, column=2, value=f'=COUNTIFS({names},$A{row})')
        sheet.cell(row=row, column=3, value=f'=SUMIFS({money_in},{names},$A{row})')
        sheet.cell(row=row, column=4, value=f'=SUMIFS({money_out},{names},$A{row})')
        sheet.cell(row=row, column=5, value=f'=SUMIFS({amounts},{names},$A{row})')
        # Share of its own direction: receipts over total receipts, or
        # payments over total payments.
        sheet.cell(row=row, column=6, value=(
            f'=IF(E{row}>=0,C{row}/SUM({money_in}),D{row}/SUM({money_out}))'
        ))
        for column in range(1, 7):
            cell = sheet.cell(row=row, column=column)
            cell.font = Font(name=FONT, size=10, color=INK)
            if index % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=BAND)
        for column in (3, 4, 5):
            sheet.cell(row=row, column=column).number_format = GBP
        sheet.cell(row=row, column=6).number_format = PCT

    total = 5 + len(companies)
    sheet.cell(row=total, column=1, value="Total").font = Font(
        name=FONT, size=10, bold=True)
    for column in (2, 3, 4, 5):
        letter = get_column_letter(column)
        cell = sheet.cell(row=total, column=column,
                          value=f"=SUM({letter}5:{letter}{total - 1})")
        cell.font = Font(name=FONT, size=10, bold=True)
        if column > 2:
            cell.number_format = GBP
    thin_rule(sheet, total, 6, "medium")
    widths(sheet, {"A": 32, "B": 14, "C": 15, "D": 15, "E": 15, "F": 18})


def build_dashboard(sheet, months: list[str], last: int) -> None:
    title_block(sheet, "Cash In / Cash Out Summary",
                "Every figure below is a formula. Change the Transactions "
                "sheet and this updates.", 4)

    monthly_last = 4 + len(months)
    figures = [
        ("Cash in", f"=SUM(Transactions!F5:F{last})", GBP, "Total receipts"),
        ("Cash out", f"=SUM(Transactions!G5:G{last})", GBP, "Total payments"),
        ("Net result", f"=SUM(Transactions!H5:H{last})", GBP,
         "Cash basis, not statutory profit"),
        ("Cost ratio", f"=SUM(Transactions!G5:G{last})/"
                       f"SUM(Transactions!F5:F{last})", PCT,
         "Spent per £1 received"),
        ("Transactions", f"=COUNT(Transactions!H5:H{last})", "0",
         "Lines in the statement"),
        ("Months covered", f"=COUNT('Monthly cash flow'!D5:D{monthly_last})",
         "0", "Periods in the data"),
        ("Missing documents", f'=COUNTBLANK(Transactions!E5:E{last})', "0",
         "Should be zero"),
        ("Best month", f"=MAX('Monthly cash flow'!D5:D{monthly_last})", GBP,
         "Largest net inflow"),
        ("Worst month", f"=MIN('Monthly cash flow'!D5:D{monthly_last})", GBP,
         "Largest net outflow"),
    ]

    header_row(sheet, 4, ["Measure", "Value", "Note"])
    for index, (label, formula, fmt, note) in enumerate(figures):
        row = 5 + index
        sheet.cell(row=row, column=1, value=label).font = Font(
            name=FONT, size=10, bold=True, color=INK)
        value = sheet.cell(row=row, column=2, value=formula)
        # Green text = a link to another sheet, the standard modelling
        # convention.
        value.font = Font(name=FONT, size=11, bold=True, color="008000")
        value.number_format = fmt
        value.alignment = Alignment(horizontal="right")
        sheet.cell(row=row, column=3, value=note).font = Font(
            name=FONT, size=9, color=MUTED)
        if index % 2 == 1:
            for column in range(1, 4):
                sheet.cell(row=row, column=column).fill = PatternFill(
                    "solid", fgColor=BAND)
        sheet.row_dimensions[row].height = 19

    guide = 5 + len(figures) + 2
    sheet.cell(row=guide, column=1, value="How to read this workbook").font = (
        Font(name=FONT, size=10, bold=True, color=INK))
    lines = [
        "Transactions - the cleaned bank statement. Columns H to K are "
        "calculated; the rest is source data.",
        "Categories - the keyword rules. Blue text on yellow means you may "
        "edit it; everything else is a formula.",
        "Monthly cash flow - SUMIFS by month, with a running total.",
        "Profit and loss - SUMIFS by category, income then expenditure.",
        "By company - what has accumulated with each counterparty.",
        "",
        "Basis: cash accounting. Receipts and payments as they cleared the "
        "bank. No accruals, prepayments, depreciation or tax.",
    ]
    for index, text in enumerate(lines):
        cell = sheet.cell(row=guide + 1 + index, column=1, value=text)
        cell.font = Font(name=FONT, size=9, color=MUTED)

    widths(sheet, {"A": 26, "B": 18, "C": 40})


# ---------------------------------------------------------------------------
def build(csv_path: Path, output: Path) -> None:
    rows = read_csv(csv_path)
    months = sorted({
        datetime.strptime(r["date"].strip(), "%d/%m/%Y").strftime("%Y-%m")
        for r in rows
    })
    companies = sorted(
        {r["counterparty"].strip() for r in rows},
        key=lambda name: -sum(
            number(r["paid in"]) + number(r["paid out"])
            for r in rows if r["counterparty"].strip() == name
        ),
    )

    workbook = Workbook()
    dashboard = workbook.active
    dashboard.title = "Summary"
    transactions = workbook.create_sheet("Transactions")
    monthly = workbook.create_sheet("Monthly cash flow")
    profit = workbook.create_sheet("Profit and loss")
    by_company = workbook.create_sheet("By company")
    categories = workbook.create_sheet("Categories")

    build_categories(categories)
    last = build_transactions(transactions, rows)
    build_monthly(monthly, months, last)
    build_profit_and_loss(profit, last)
    build_by_company(by_company, companies, last)
    build_dashboard(dashboard, months, last)

    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False
        printable(sheet)

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    print(f"Written {output}")
    print(f"  {len(rows)} transactions, {len(months)} months, "
          f"{len(companies)} companies")


def main(argv: list[str]) -> int:
    csv_path = Path(argv[1]) if len(argv) > 1 else DEFAULT_CSV
    if not csv_path.exists():
        print(f"CSV not found: {csv_path}", file=sys.stderr)
        return 1
    build(csv_path, OUTPUT)
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
