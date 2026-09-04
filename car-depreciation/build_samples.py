"""
Build calibrated used-market samples for the Vauxhall Astra and VW Golf.

The prices are GENERATED, not scraped: UK listing sites are unreachable from
the environment this was written in. Each car's curve is calibrated against
published UK depreciation figures, recorded with sources on the Calibration
tab of the workbook.

The Astra config and its RNG call order are unchanged from the original
single-car generator, so `data/vauxhall_astra_market_sample.csv` reproduces
byte for byte.
"""

import csv
from pathlib import Path

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR = Path(__file__).resolve().parent / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)
VALUATION_YEAR = 2026

REGIONS = ["South East", "North West", "West Midlands", "Yorkshire", "Scotland",
           "South West", "East of England", "Wales", "London", "North East"]

# Shared non-age effects, so any fitted difference between the two cars comes
# from the age curve rather than from differently-rigged extras.
FUEL_FACTOR = {"Petrol": 1.00, "Diesel": 0.97, "Hybrid": 1.06,
               "Plug-in Hybrid": 1.02, "Electric": 0.88}
TRANS_FACTOR = {"Manual": 1.00, "Automatic": 1.04}
HISTORY_FACTOR = {"Full": 1.00, "Partial": 0.94, "None": 0.87}
SELLER_FACTOR = {"Franchise dealer": 1.08, "Independent dealer": 1.00, "Private": 0.90}
COND_FACTOR = {"Excellent": 1.04, "Good": 1.00, "Fair": 0.92}
TRIM_FACTOR = {
    # Astra
    "Design": 0.95, "SRi": 1.02, "SRi VX-Line": 1.05, "Elite Nav": 1.08,
    "GS Line": 1.03, "GS": 1.05, "Ultimate": 1.10, "Griffin": 0.97,
    # Golf
    "S": 0.94, "Match": 0.98, "GT": 1.03, "R-Line": 1.06, "GTI": 1.18,
    "Life": 0.95, "Style": 1.04, "Black Edition": 1.09,
}

# ---------------------------------------------------------------------------
# Per-car calibration
# ---------------------------------------------------------------------------
ASTRA = {
    "key": "astra",
    "name": "Vauxhall Astra",
    "file": "vauxhall_astra_market_sample",
    "seed": 20260826,
    "list_by_year": {2026: 29995, 2025: 29500, 2024: 28500, 2023: 27500, 2022: 25500,
                     2021: 23000, 2020: 22000, 2019: 21000, 2018: 20000,
                     2017: 19500, 2016: 19000},
    # ~50% retained at 3yr, ~42% at 5yr; flattens from year 7-8.
    "retention": {0: 0.92, 1: 0.74, 2: 0.62, 3: 0.51, 4: 0.455, 5: 0.42,
                  6: 0.375, 7: 0.335, 8: 0.30, 9: 0.275, 10: 0.255},
    "anchors": [
        ("2026 list price, entry 1.2 Turbo petrol", "£27,495", "Stellantis Media / Auto Express"),
        ("Retained at 3 years", "47-53%", "CarBuyerIQ / Free Plate Check"),
        ("Retained at 5 years", "~42%", "Free Plate Check"),
        ("Depreciation flattens from", "year 7-9", "CarBuyerIQ"),
        ("Class position", "among the worst in the family hatchback class", "CarBuyerIQ"),
    ],
}

GOLF = {
    "key": "golf",
    "name": "Volkswagen Golf",
    "file": "vw_golf_market_sample",
    "seed": 20260901,
    "list_by_year": {2026: 30500, 2025: 30000, 2024: 29000, 2023: 28000, 2022: 26500,
                     2021: 25000, 2020: 24000, 2019: 23000, 2018: 22500,
                     2017: 22000, 2016: 21500},
    # 61% retained at 3yr, 42-48% at 5yr; keeps depreciating to 10-12 years,
    # so the tail stays steeper than the Astra's rather than levelling off.
    "retention": {0: 0.93, 1: 0.80, 2: 0.70, 3: 0.61, 4: 0.53, 5: 0.45,
                  6: 0.405, 7: 0.365, 8: 0.33, 9: 0.30, 10: 0.275},
    "anchors": [
        ("2026 on-the-road price, from", "£28,910", "heycar / Carwow"),
        ("Retained at 3 years", "~61%", "CarBuyerIQ"),
        ("Retained at 5 years", "42-48%", "CarBuyerIQ"),
        ("Keeps depreciating until", "year 10-12", "CarBuyerIQ"),
        ("Class position", "second only to the Toyota Corolla in the UK", "CarBuyerIQ"),
    ],
}
CARS = [ASTRA, GOLF]

# Identical mileage penalty for both cars, so a fitted difference in b is noise
# rather than something baked in.
MILEAGE_PENALTY = 0.045
EXPECTED_MILES_PER_YEAR = 10_000


def _trim_and_fuel(cfg, rng, reg_year, pick):
    """Trim and fuel for a registration year. Order of rng calls is fixed."""
    if cfg["key"] == "astra":
        is_k = reg_year in range(2016, 2022)              # previous generation
        trim = (pick(["Design", "SRi", "SRi VX-Line", "Elite Nav", "Ultimate"]) if is_k
                else pick(["Design", "GS Line", "GS", "Ultimate", "Griffin"]
                          if reg_year < 2026 else ["Griffin", "GS", "Ultimate"]))
        if is_k:
            fuel = pick(["Petrol", "Diesel"], p=[0.72, 0.28])
        elif reg_year == 2022:
            fuel = pick(["Petrol", "Diesel", "Plug-in Hybrid"], p=[0.6, 0.2, 0.2])
        else:
            fuel = pick(["Petrol", "Hybrid", "Plug-in Hybrid", "Electric"],
                        p=[0.42, 0.28, 0.16, 0.14])
        return trim, fuel

    # Golf: Mk7/7.5 to 2019, Mk8 from 2020. No full EV — that is the ID.3.
    is_mk7 = reg_year <= 2019
    trim = (pick(["S", "Match", "GT", "R-Line", "GTI"]) if is_mk7
            else pick(["Life", "Match", "Style", "R-Line", "Black Edition", "GTI"]))
    if is_mk7:
        fuel = pick(["Petrol", "Diesel"], p=[0.63, 0.37])
    elif reg_year <= 2021:
        fuel = pick(["Petrol", "Diesel", "Hybrid", "Plug-in Hybrid"],
                    p=[0.5, 0.22, 0.18, 0.10])
    else:
        fuel = pick(["Petrol", "Hybrid", "Plug-in Hybrid"], p=[0.52, 0.32, 0.16])
    return trim, fuel


def generate(cfg, n=100):
    """Generate one car's sample. Deterministic given cfg['seed']."""
    rng = np.random.default_rng(cfg["seed"])

    def pick(options, p=None):
        return str(rng.choice(options, p=p))

    # Weighted toward the 2-7 year band that dominates real forecourt stock.
    weights = np.array([3, 7, 11, 13, 13, 12, 11, 9, 8, 7, 6], dtype=float)
    weights /= weights.sum()
    ages = rng.choice(np.arange(0, 11), size=n, p=weights)
    ages.sort()

    rows = []
    for i, age in enumerate(ages, start=1):
        age = int(age)
        reg_year = VALUATION_YEAR - age
        list_price = cfg["list_by_year"][reg_year]

        # Mileage drawn independently of age, so the two are not collinear.
        annual = rng.uniform(3500, 21000)
        mileage = max(int(annual * age + rng.normal(0, 1500)), 25 if age == 0 else 400)

        trim, fuel = _trim_and_fuel(cfg, rng, reg_year, pick)
        transmission = ("Automatic" if fuel in ("Electric", "Hybrid", "Plug-in Hybrid")
                        else pick(["Manual", "Automatic"], p=[0.62, 0.38]))
        engine = {"Electric": "-", "Hybrid": "1.2", "Plug-in Hybrid": "1.6"}.get(
            fuel, pick(["1.2", "1.4", "1.5", "1.6"]))
        body = pick(["Hatchback", "Sports Tourer"], p=[0.78, 0.22])
        owners = int(min(1 + rng.poisson(max(age - 1, 0) * 0.45), 6)) if age else 1
        history = "Full" if age <= 2 else pick(["Full", "Partial", "None"], p=[0.62, 0.29, 0.09])
        seller = pick(["Franchise dealer", "Independent dealer", "Private"], p=[0.34, 0.46, 0.20])
        cond = pick(["Excellent", "Good", "Fair"], p=[0.22, 0.62, 0.16])

        expected = EXPECTED_MILES_PER_YEAR * age
        mileage_factor = float(np.clip(
            np.exp(-MILEAGE_PENALTY * (mileage - expected) / 10_000), 0.72, 1.22))
        owner_factor = 1.0 - 0.015 * max(owners - 1, 0)

        price = (list_price * cfg["retention"][age] * mileage_factor * TRIM_FACTOR[trim]
                 * FUEL_FACTOR[fuel] * TRANS_FACTOR[transmission] * HISTORY_FACTOR[history]
                 * SELLER_FACTOR[seller] * COND_FACTOR[cond] * owner_factor
                 * float(rng.normal(1.0, 0.045)))
        price = int(round(max(price, 900) / 5) * 5)      # forecourt prices land on £5

        rows.append([f"{cfg['key'][:3].upper()}-{i:03d}", reg_year, mileage, trim, fuel,
                     transmission, engine, body, owners, history, seller, cond,
                     pick(REGIONS), list_price, price])
    return rows


HEADER = ["listing_id", "reg_year", "age_years", "mileage", "avg_annual_mileage",
          "trim", "fuel", "transmission", "engine_litres", "body", "former_keepers",
          "service_history", "seller_type", "condition", "region",
          "original_list_gbp", "asking_price_gbp", "retained_pct"]


def to_records(cfg, rows):
    out = []
    for (lid, reg, mil, trim, fuel, trans, eng, body, own, hist, sell, cond,
         region, listp, price) in rows:
        age = VALUATION_YEAR - reg
        out.append([lid, reg, age, mil, "" if age == 0 else round(mil / age),
                    trim, fuel, trans, eng, body, own, hist, sell, cond, region,
                    listp, price, round(price / listp, 4)])
    return out


def write_csv(path, records, model_name=None):
    with open(path, "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow((["model"] if model_name is not None else []) + HEADER)
        for r in records:
            w.writerow(([model_name] if model_name is not None else []) + r)


# ---------------------------------------------------------------------------
# Workbook
# ---------------------------------------------------------------------------
ARIAL = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
BODY = Font(name=ARIAL, size=10)
BLUE = Font(name=ARIAL, size=10, color="0000FF")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _sheet_of_listings(wb, title, records):
    ws = wb.create_sheet(title)
    headings = ["Listing ID", "Registration Year", "Age (years)", "Mileage",
                "Avg Annual Mileage", "Trim", "Fuel", "Transmission", "Engine (L)",
                "Body", "Former Keepers", "Service History", "Seller Type",
                "Condition", "Region", "Original List Price (GBP)",
                "Asking Price (GBP)", "Retained % of List"]
    for col, h in enumerate(headings, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill = HDR_FONT, HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.freeze_panes = "A2"
    for i, rec in enumerate(records, start=2):
        for col, v in enumerate(rec, start=1):
            c = ws.cell(row=i, column=col, value=v)
            c.font, c.border = BODY, BORDER
            if col in (4, 5, 16, 17):
                c.number_format = '#,##0'
            if col == 18:
                c.number_format = '0.0%'
    for col, w in enumerate([11, 12, 9, 10, 12, 14, 14, 13, 10, 14, 11, 13, 18, 11,
                             15, 15, 14, 11], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    return ws, len(records) + 1


def build_workbook(per_car):
    wb = Workbook()
    ws = wb.active
    ws.title = "Read Me"
    lines = [
        ("Used Market Samples — Vauxhall Astra vs Volkswagen Golf", 14, True, "C00000"),
        ("", 10, False, None),
        ("WHAT THIS IS", 11, True, "1F3864"),
        ("200 cars, 100 of each model, aged 0-10, generated from a documented price model", 10, False, None),
        ("calibrated against published UK depreciation figures. See the Calibration tab.", 10, False, None),
        ("", 10, False, None),
        ("WHAT THIS IS NOT", 11, True, "C00000"),
        ("These are NOT real listings. No row is a real car, advert, seller or transaction.", 10, False, "C00000"),
        ("Prices were computed, not observed. UK listing sites were unreachable from the", 10, False, "C00000"),
        ("environment this was built in, so no adverts could be collected. Do not quote any", 10, False, "C00000"),
        ("figure here as evidence of what the market actually did.", 10, False, "C00000"),
        ("", 10, False, None),
        ("WHAT IS REAL", 11, True, "1F3864"),
        ("The calibration anchors are real published figures with sources, and the gap they", 10, False, None),
        ("describe is real: the Golf retains ~61% of its value at three years against the", 10, False, None),
        ("Astra's 47-53%, and the Golf keeps depreciating to 10-12 years where the Astra", 10, False, None),
        ("flattens at 7-9. The samples reproduce that gap; they do not discover it.", 10, False, None),
        ("", 10, False, None),
        ("HOW THE TWO ARE MADE COMPARABLE", 11, True, "1F3864"),
        ("Mileage penalty, trim, fuel, history, seller and condition multipliers are IDENTICAL", 10, False, None),
        ("for both cars. Only the list prices and the retention curves differ, so any fitted", 10, False, None),
        ("difference other than the age curve is noise rather than something rigged in.", 10, False, None),
        ("", 10, False, None),
        ("BEFORE SHOWING THIS TO AN EMPLOYER", 11, True, "1F3864"),
        ("Replace it with real data. Kaggle's '100,000 UK Used Car Data set' has real scraped", 10, False, None),
        ("listings including both vauxhall.csv and vw.csv — the same comparison, done for real.", 10, False, None),
    ]
    for r, (text, size, bold, color) in enumerate(lines, start=1):
        ws.cell(row=r, column=1, value=text).font = Font(
            name=ARIAL, size=size, bold=bold, color=color or "000000")
    ws.column_dimensions["A"].width = 100

    ranges = {}
    for cfg, records in per_car:
        _, n = _sheet_of_listings(wb, f"{cfg['name'].split()[-1]} Listings", records)
        ranges[cfg["key"]] = (f"{cfg['name'].split()[-1]} Listings", n)

    # ---- Comparison -------------------------------------------------------
    ws = wb.create_sheet("Comparison")
    ws["A1"] = "Mean asking price by age"
    ws["A1"].font = Font(name=ARIAL, bold=True, size=12, color="1F3864")
    ws["A2"] = "Live formulas over the two listings tabs."
    ws["A2"].font = Font(name=ARIAL, size=9, italic=True)
    heads = ["Age (years)", "Astra cars", "Astra mean price", "Golf cars",
             "Golf mean price", "Golf premium (£)", "Golf premium (%)"]
    for col, h in enumerate(heads, start=1):
        c = ws.cell(row=4, column=col, value=h)
        c.font, c.fill = HDR_FONT, HDR_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = BORDER
    a_sheet, a_n = ranges["astra"]
    g_sheet, g_n = ranges["golf"]
    for k, age in enumerate(range(0, 11)):
        r = 5 + k
        ws.cell(row=r, column=1, value=age)
        ws.cell(row=r, column=2, value=f"=COUNTIFS('{a_sheet}'!$C$2:$C${a_n},A{r})")
        ws.cell(row=r, column=3, value=f"=IF(B{r}=0,\"\",ROUND(AVERAGEIFS('{a_sheet}'!$Q$2:$Q${a_n},'{a_sheet}'!$C$2:$C${a_n},A{r}),0))")
        ws.cell(row=r, column=4, value=f"=COUNTIFS('{g_sheet}'!$C$2:$C${g_n},A{r})")
        ws.cell(row=r, column=5, value=f"=IF(D{r}=0,\"\",ROUND(AVERAGEIFS('{g_sheet}'!$Q$2:$Q${g_n},'{g_sheet}'!$C$2:$C${g_n},A{r}),0))")
        ws.cell(row=r, column=6, value=f'=IF(OR(C{r}="",E{r}=""),"",E{r}-C{r})')
        ws.cell(row=r, column=7, value=f'=IF(OR(C{r}="",E{r}="",C{r}=0),"",E{r}/C{r}-1)')
        for col in range(1, 8):
            cell = ws.cell(row=r, column=col)
            cell.font, cell.border = BODY, BORDER
            if col in (3, 5, 6):
                cell.number_format = '£#,##0'
            if col == 7:
                cell.number_format = '0.0%'
            if col in (1, 2, 4):
                cell.alignment = Alignment(horizontal="center")
    for col, w in zip("ABCDEFG", (12, 11, 18, 11, 18, 17, 17)):
        ws.column_dimensions[col].width = w
    note = [
        "Reading the table",
        "The Golf premium is widest in the middle years and narrows at both ends: a nearly-new",
        "car of either badge is close to its list price, and by ten years both are cheap enough",
        "that the gap in pounds is small. The percentage column is the one that matters.",
        "",
        "Some age buckets hold only a handful of cars, so one unusual car moves that row.",
        "Read the shape of the column, not any single cell.",
    ]
    for k, t in enumerate(note):
        ws.cell(row=17 + k, column=1, value=t).font = Font(
            name=ARIAL, size=10, bold=(k == 0), color="1F3864" if k == 0 else "000000")

    # ---- Calibration ------------------------------------------------------
    ws = wb.create_sheet("Calibration")
    ws["A1"] = "Calibration anchors, sources and generating parameters"
    ws["A1"].font = Font(name=ARIAL, bold=True, size=12, color="1F3864")
    ws["A3"], ws["B3"] = "Valuation year", VALUATION_YEAR
    ws["A3"].font = Font(name=ARIAL, bold=True, size=10)
    ws["B3"].font = BLUE

    row = 5
    for cfg, _ in per_car:
        ws.cell(row=row, column=1, value=f"{cfg['name']} — published anchors").font = Font(
            name=ARIAL, bold=True, size=11, color="1F3864")
        row += 1
        for col, h in zip("ABC", ("Anchor", "Value", "Source")):
            c = ws[f"{col}{row}"]
            c.value, c.font, c.fill, c.border = h, HDR_FONT, HDR_FILL, BORDER
        row += 1
        for a, v, s in cfg["anchors"]:
            for col, val in zip("ABC", (a, v, s)):
                c = ws[f"{col}{row}"]
                c.value, c.font, c.border = val, BODY, BORDER
            row += 1
        row += 1

    ws.cell(row=row, column=1, value="Retention curves (share of original list retained)").font = Font(
        name=ARIAL, bold=True, size=11, color="1F3864")
    row += 1
    for col, h in zip("ABC", ("Age (years)", "Astra", "Golf")):
        c = ws[f"{col}{row}"]
        c.value, c.font, c.fill, c.border = h, HDR_FONT, HDR_FILL, BORDER
    row += 1
    for age in range(0, 11):
        ws.cell(row=row, column=1, value=age).font = BODY
        for col, cfg in ((2, ASTRA), (3, GOLF)):
            c = ws.cell(row=row, column=col, value=cfg["retention"][age])
            c.font, c.number_format = BLUE, '0.0%'
        for col in (1, 2, 3):
            ws.cell(row=row, column=col).border = BORDER
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Shared generating parameters (identical for both cars)").font = Font(
        name=ARIAL, bold=True, size=11, color="1F3864")
    row += 1
    for a, b in [
        ("Mileage penalty", f"exp(-{MILEAGE_PENALTY} x excess per 10k miles), clipped to [0.72, 1.22]"),
        ("Expected mileage for age", f"{EXPECTED_MILES_PER_YEAR:,} miles per year"),
        ("Annual mileage drawn from", "Uniform(3,500, 21,000) — independent of age"),
        ("Price noise", "Normal(mean 1.0, sd 4.5%)"),
        ("Fuel multiplier", "0.88 (Electric) to 1.06 (Hybrid)"),
        ("Service history multiplier", "1.00 Full / 0.94 Partial / 0.87 None"),
        ("Seller multiplier", "1.08 Franchise / 1.00 Independent / 0.90 Private"),
        ("Condition multiplier", "1.04 Excellent / 1.00 Good / 0.92 Fair"),
        ("Former keeper penalty", "1.5% per keeper beyond the first"),
        ("Seeds", f"Astra {ASTRA['seed']}, Golf {GOLF['seed']} — regenerating reproduces both"),
    ]:
        ws.cell(row=row, column=1, value=a).font = Font(name=ARIAL, size=10, bold=True)
        ws.cell(row=row, column=2, value=b).font = BODY
        row += 1

    for col, w in zip("ABCDE", (36, 34, 46, 20, 20)):
        ws.column_dimensions[col].width = w
    return wb


if __name__ == "__main__":
    per_car = []
    for cfg in CARS:
        rows = generate(cfg)
        records = to_records(cfg, rows)
        write_csv(DATA_DIR / f"{cfg['file']}.csv", records)
        per_car.append((cfg, records))
        prices = [r[16] for r in records]
        print(f"{cfg['name']:<18} {len(records)} cars | "
              f"£{min(prices):,}-£{max(prices):,} | "
              f"mean £{sum(prices) / len(prices):,.0f}")

    combined = DATA_DIR / "combined_market_sample.csv"
    with open(combined, "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["model"] + HEADER)
        for cfg, records in per_car:
            for r in records:
                w.writerow([cfg["name"]] + r)
    print("wrote", combined)

    wb = build_workbook(per_car)
    out = DATA_DIR / "astra_vs_golf_market_sample.xlsx"
    wb.save(out)
    print("wrote", out)
