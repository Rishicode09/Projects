"""Build a 100-row synthetic-but-calibrated Vauxhall Astra used-market sample."""
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from pathlib import Path

# Written next to this script, so the generator works from any checkout rather
# than only on the machine it was first written on.
DATA_DIR = Path(__file__).resolve().parent / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)
OUT = str(DATA_DIR / "vauxhall_astra_market_sample.xlsx")
VALUATION_YEAR = 2026
rng = np.random.default_rng(20260826)

# --- Calibration anchors (see Calibration sheet for sources) --------------
# Original UK list price (approx, entry-to-mid trim) by registration year.
LIST_BY_YEAR = {
    2026: 29995, 2025: 29500, 2024: 28500, 2023: 27500, 2022: 25500,
    2021: 23000, 2020: 22000, 2019: 21000, 2018: 20000, 2017: 19500, 2016: 19000,
}
# Share of original list retained, by age. Anchored on ~50% at 3yr and ~42%
# at 5yr, flattening after year 7.
RETENTION = {
    0: 0.92, 1: 0.74, 2: 0.62, 3: 0.51, 4: 0.455, 5: 0.42,
    6: 0.375, 7: 0.335, 8: 0.30, 9: 0.275, 10: 0.255,
}

ASTRA_K = range(2016, 2022)          # previous generation
TRIMS_K = ["Design", "SRi", "SRi VX-Line", "Elite Nav", "Ultimate"]
TRIMS_L = ["Design", "GS Line", "GS", "Ultimate", "Griffin"]
TRIM_FACTOR = {
    "Design": 0.95, "SRi": 1.02, "SRi VX-Line": 1.05, "Elite Nav": 1.08,
    "GS Line": 1.03, "GS": 1.05, "Ultimate": 1.10, "Griffin": 0.97,
}
FUEL_FACTOR = {"Petrol": 1.00, "Diesel": 0.97, "Hybrid": 1.06,
               "Plug-in Hybrid": 1.02, "Electric": 0.88}
TRANS_FACTOR = {"Manual": 1.00, "Automatic": 1.04}
HISTORY_FACTOR = {"Full": 1.00, "Partial": 0.94, "None": 0.87}
SELLER_FACTOR = {"Franchise dealer": 1.08, "Independent dealer": 1.00, "Private": 0.90}
COND_FACTOR = {"Excellent": 1.04, "Good": 1.00, "Fair": 0.92}
REGIONS = ["South East", "North West", "West Midlands", "Yorkshire", "Scotland",
           "South West", "East of England", "Wales", "London", "North East"]

def pick(options, p=None):
    return str(rng.choice(options, p=p))

rows = []
# Spread 100 cars across ages 0-10, weighted toward the 2-7 year band that
# dominates real forecourt stock.
age_weights = np.array([3, 7, 11, 13, 13, 12, 11, 9, 8, 7, 6], dtype=float)
age_weights /= age_weights.sum()
ages = rng.choice(np.arange(0, 11), size=100, p=age_weights)
ages.sort()

for i, age in enumerate(ages, start=1):
    age = int(age)
    reg_year = VALUATION_YEAR - age
    list_price = LIST_BY_YEAR[reg_year]
    is_k = reg_year in ASTRA_K

    # Mileage is drawn independently of age, so the two are NOT collinear.
    annual = rng.uniform(3500, 21000)
    mileage = max(int(annual * age + rng.normal(0, 1500)), 25 if age == 0 else 400)

    trim = pick(TRIMS_K) if is_k else pick(TRIMS_L if reg_year < 2026 else ["Griffin", "GS", "Ultimate"])
    if is_k:
        fuel = pick(["Petrol", "Diesel"], p=[0.72, 0.28])
    elif reg_year == 2022:
        fuel = pick(["Petrol", "Diesel", "Plug-in Hybrid"], p=[0.6, 0.2, 0.2])
    else:
        fuel = pick(["Petrol", "Hybrid", "Plug-in Hybrid", "Electric"], p=[0.42, 0.28, 0.16, 0.14])

    # Astra electrified variants are all auto-only; only petrol/diesel offer a manual.
    transmission = ("Automatic" if fuel in ("Electric", "Hybrid", "Plug-in Hybrid")
                    else pick(["Manual", "Automatic"], p=[0.62, 0.38]))
    engine = {"Electric": "-", "Hybrid": "1.2", "Plug-in Hybrid": "1.6"}.get(fuel, pick(["1.2", "1.4", "1.5", "1.6"]))
    body = pick(["Hatchback", "Sports Tourer"], p=[0.78, 0.22])
    owners = int(min(1 + rng.poisson(max(age - 1, 0) * 0.45), 6)) if age else 1
    history = "Full" if age <= 2 else pick(["Full", "Partial", "None"], p=[0.62, 0.29, 0.09])
    seller = pick(["Franchise dealer", "Independent dealer", "Private"], p=[0.34, 0.46, 0.20])
    cond = pick(["Excellent", "Good", "Fair"], p=[0.22, 0.62, 0.16])

    # Price model. Mileage is penalised relative to what is normal for the age,
    # so a low-mileage old car and a high-mileage new one both price sensibly.
    expected_mileage = 10000 * age
    mileage_factor = float(np.exp(-0.045 * (mileage - expected_mileage) / 10000))
    mileage_factor = float(np.clip(mileage_factor, 0.72, 1.22))
    owner_factor = 1.0 - 0.015 * max(owners - 1, 0)

    price = (list_price * RETENTION[age] * mileage_factor * TRIM_FACTOR[trim]
             * FUEL_FACTOR[fuel] * TRANS_FACTOR[transmission] * HISTORY_FACTOR[history]
             * SELLER_FACTOR[seller] * COND_FACTOR[cond] * owner_factor
             * float(rng.normal(1.0, 0.045)))
    price = int(round(max(price, 900) / 5) * 5)   # forecourt prices land on £5

    rows.append([f"AST-{i:03d}", reg_year, mileage, trim, fuel, transmission,
                 engine, body, owners, history, seller, cond,
                 pick(REGIONS), list_price, price])

# ---------------------------------------------------------------- workbook
wb = Workbook()
ARIAL = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
BLUE = Font(name=ARIAL, size=10, color="0000FF")
BODY = Font(name=ARIAL, size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ---- Read Me -------------------------------------------------------------
ws = wb.active
ws.title = "Read Me"
readme = [
    ("Vauxhall Astra — Used Market Sample (100 cars)", 14, True, "C00000"),
    ("", 10, False, None),
    ("WHAT THIS IS", 11, True, "1F3864"),
    ("A 100-row sample of the UK used Vauxhall Astra market, generated from a documented", 10, False, None),
    ("price model calibrated against published depreciation figures. See the Calibration tab.", 10, False, None),
    ("", 10, False, None),
    ("WHAT THIS IS NOT", 11, True, "C00000"),
    ("These are NOT real listings. No row corresponds to a real car, a real advert, a real", 10, False, "C00000"),
    ("seller, or a real transaction. Prices were computed, not observed. Do not quote any", 10, False, "C00000"),
    ("figure here as evidence of what the market actually did.", 10, False, "C00000"),
    ("", 10, False, None),
    ("WHY IT EXISTS", 11, True, "1F3864"),
    ("The depreciation model needs data where age and mileage vary independently. In the", 10, False, None),
    ("original synthetic data, mileage was ~1.1 x age (correlation 0.999), so the model could", 10, False, None),
    ("not separate the age effect from the mileage effect and the mileage coefficient pinned", 10, False, None),
    ("to its bound in 85% of runs. Here, annual mileage is drawn independently between 3,500", 10, False, None),
    ("and 21,000 miles, so the two effects are genuinely separable.", 10, False, None),
    ("", 10, False, None),
    ("HOW TO USE IT", 11, True, "1F3864"),
    ("Fit on columns: age (from Registration Year), Mileage, and Asking Price. The generating", 10, False, None),
    ("parameters are listed on the Calibration tab, so you can check whether your fitted", 10, False, None),
    ("parameters recover them. That is the correct use of synthetic data: a test that your", 10, False, None),
    ("estimator works, before you point it at real listings.", 10, False, None),
    ("", 10, False, None),
    ("BEFORE SHOWING THIS TO AN EMPLOYER", 11, True, "1F3864"),
    ("Replace it with real data. Kaggle's '100,000 UK Used Car Data set' has real scraped", 10, False, None),
    ("listings including a vauxhall.csv. Real data is the thing that makes the project count;", 10, False, None),
    ("this file only proves the pipeline runs.", 10, False, None),
]
for r, (text, size, bold, color) in enumerate(readme, start=1):
    c = ws.cell(row=r, column=1, value=text)
    c.font = Font(name=ARIAL, size=size, bold=bold, color=color or "000000")
ws.column_dimensions["A"].width = 100

# ---- Listings ------------------------------------------------------------
ws = wb.create_sheet("Listings")
headers = ["Listing ID", "Registration Year", "Age (years)", "Mileage",
           "Avg Annual Mileage", "Trim", "Fuel", "Transmission", "Engine (L)",
           "Body", "Former Keepers", "Service History", "Seller Type",
           "Condition", "Region", "Original List Price (GBP)",
           "Asking Price (GBP)", "Retained % of List"]
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=1, column=col, value=h)
    c.font, c.fill = HDR_FONT, HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
ws.freeze_panes = "A2"

for i, r in enumerate(rows, start=2):
    (lid, reg, mil, trim, fuel, trans, eng, body, own, hist, sell, cond, region,
     listp, price) = r
    vals = [lid, reg, f"=Calibration!$B$3-B{i}", mil,
            f'=IF(C{i}=0,"",ROUND(D{i}/C{i},0))', trim, fuel, trans, eng, body,
            own, hist, sell, cond, region, listp, price, f"=Q{i}/P{i}"]
    for col, v in enumerate(vals, start=1):
        c = ws.cell(row=i, column=col, value=v)
        c.font = BODY
        c.border = BORDER
        if col in (4, 5, 16, 17):
            c.number_format = '#,##0'
        if col == 18:
            c.number_format = '0.0%'
        if col in (2, 3, 11):
            c.alignment = Alignment(horizontal="center")

widths = [11, 12, 9, 10, 12, 13, 14, 13, 10, 14, 11, 13, 18, 11, 15, 15, 14, 11]
for col, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(col)].width = w

n = len(rows) + 1
# Totals strip
tr = n + 2
ws.cell(row=tr, column=1, value="Sample summary").font = Font(name=ARIAL, bold=True, size=10)
for label, formula, fmt, col in [
    ("Cars", f"=COUNT(Q2:Q{n})", '#,##0', 3),
    ("Mean asking price", f"=ROUND(AVERAGE(Q2:Q{n}),0)", '£#,##0', 5),
    ("Mean mileage", f"=ROUND(AVERAGE(D2:D{n}),0)", '#,##0', 7),
    ("Mean age", f"=ROUND(AVERAGE(C2:C{n}),1)", '0.0', 9),
]:
    ws.cell(row=tr, column=col - 1, value=label).font = Font(name=ARIAL, bold=True, size=10)
    c = ws.cell(row=tr, column=col, value=formula)
    c.font, c.number_format = Font(name=ARIAL, bold=True, size=10), fmt

# ---- Market Summary ------------------------------------------------------
ws = wb.create_sheet("Market Summary")
ws["A1"] = "How the sample prices by age"
ws["A1"].font = Font(name=ARIAL, bold=True, size=12, color="1F3864")
ws["A2"] = "Every figure below is a live formula over the Listings tab."
ws["A2"].font = Font(name=ARIAL, size=9, italic=True)

sum_hdr = ["Age (years)", "Cars", "Mean Asking Price", "Mean Mileage",
           "Mean Retained % of List", "Mean Price per 1k Miles"]
for col, h in enumerate(sum_hdr, start=1):
    c = ws.cell(row=4, column=col, value=h)
    c.font, c.fill = HDR_FONT, HDR_FILL
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    c.border = BORDER

for k, age in enumerate(range(0, 11)):
    r = 5 + k
    ws.cell(row=r, column=1, value=age).font = BODY
    ws.cell(row=r, column=2, value=f"=COUNTIFS(Listings!$C$2:$C${n},A{r})")
    ws.cell(row=r, column=3, value=f'=IF(B{r}=0,"",ROUND(AVERAGEIFS(Listings!$Q$2:$Q${n},Listings!$C$2:$C${n},A{r}),0))')
    ws.cell(row=r, column=4, value=f'=IF(B{r}=0,"",ROUND(AVERAGEIFS(Listings!$D$2:$D${n},Listings!$C$2:$C${n},A{r}),0))')
    ws.cell(row=r, column=5, value=f'=IF(B{r}=0,"",AVERAGEIFS(Listings!$R$2:$R${n},Listings!$C$2:$C${n},A{r}))')
    ws.cell(row=r, column=6, value=f'=IF(OR(B{r}=0,D{r}=0),"",ROUND(C{r}/(D{r}/1000),0))')
    for col in range(1, 7):
        cell = ws.cell(row=r, column=col)
        cell.font, cell.border = BODY, BORDER
        if col in (3, 4):
            cell.number_format = '#,##0'
        if col == 5:
            cell.number_format = '0.0%'
        if col == 6:
            cell.number_format = '£#,##0'
        if col in (1, 2):
            cell.alignment = Alignment(horizontal="center")

for col, w in enumerate([12, 8, 18, 14, 20, 20], start=1):
    ws.column_dimensions[get_column_letter(col)].width = w

r0 = 17
notes = [
    "Reading the table",
    "The steep early drop then the flattening tail is the two-phase shape the model fits:",
    "a fast rate for the first few years, a slower one after. 'Mean Price per 1k Miles' is a",
    "sanity check, not a valuation - it rises for young cars simply because they have few miles.",
    "",
    "Caution: with 100 cars spread over 11 ages, some age buckets hold only a handful of cars,",
    "so a single unusual car moves that row's mean. Read the shape of the column, not any one cell.",
]
for k, t in enumerate(notes):
    c = ws.cell(row=r0 + k, column=1, value=t)
    c.font = Font(name=ARIAL, size=10, bold=(k == 0), color="1F3864" if k == 0 else "000000")

# ---- Calibration ---------------------------------------------------------
ws = wb.create_sheet("Calibration")
ws["A1"] = "Calibration anchors and generating parameters"
ws["A1"].font = Font(name=ARIAL, bold=True, size=12, color="1F3864")
ws["A3"] = "Valuation year"
ws["A3"].font = Font(name=ARIAL, bold=True, size=10)
ws["B3"] = VALUATION_YEAR
ws["B3"].font = BLUE
ws["C3"] = "Blue cells are inputs. Age on the Listings tab is derived from this."
ws["C3"].font = Font(name=ARIAL, size=9, italic=True)

ws["A5"] = "Published figures used as anchors"
ws["A5"].font = Font(name=ARIAL, bold=True, size=11, color="1F3864")
anchors = [
    ("2026 Astra list price, entry 1.2 Turbo petrol", "£27,495", "Stellantis Media / Auto Express"),
    ("2026 Astra list price, Griffin / GS / Ultimate", "£29,995 / £31,495 / £33,995", "Stellantis Media"),
    ("Share of value retained at 3 years", "~48-52%", "Free Plate Check / CarBuyerIQ"),
    ("Share of value retained at 5 years", "~42%", "Free Plate Check"),
    ("Depreciation flattens from", "year 7-8", "CarBuyerIQ"),
    ("Reported average used Astra asking price", "£9,117", "AutoUncle (all model years)"),
    ("Reported average used Astra mileage", "22,623", "AutoUncle"),
]
ws["A6"], ws["B6"], ws["C6"] = "Anchor", "Value", "Source"
for col in "ABC":
    ws[f"{col}6"].font, ws[f"{col}6"].fill = HDR_FONT, HDR_FILL
    ws[f"{col}6"].border = BORDER
for k, (a, v, s) in enumerate(anchors):
    r = 7 + k
    for col, val in zip("ABC", (a, v, s)):
        c = ws[f"{col}{r}"]
        c.value, c.font, c.border = val, BODY, BORDER

ws["A16"] = "Conflicting sources - resolved as follows"
ws["A16"].font = Font(name=ARIAL, bold=True, size=11, color="C00000")
conflict = [
    "Search results also reported 2019 Astras at £10,000-£12,000 and 2020 at £13,000-£15,000.",
    "Those figures imply ~52% retained at seven years, which contradicts the ~42%-at-five-years",
    "anchor from the same search. The two cannot both hold. This file follows the retention-curve",
    "anchors, which were reported consistently across sources, and treats the listing snippets as",
    "unreliable - they likely mix trims, generations and conditions. A 2019 Astra here prices at",
    "roughly £7,000, not £11,000. If real data says otherwise, the real data wins.",
]
for k, t in enumerate(conflict):
    ws.cell(row=17 + k, column=1, value=t).font = Font(name=ARIAL, size=10)

ws["A25"] = "Generating parameters - fit the model and check you recover these"
ws["A25"].font = Font(name=ARIAL, bold=True, size=11, color="1F3864")
ws["A26"], ws["B26"] = "Age (years)", "Retained share of original list"
for col in "AB":
    ws[f"{col}26"].font, ws[f"{col}26"].fill = HDR_FONT, HDR_FILL
    ws[f"{col}26"].border = BORDER
for k, (age, ret) in enumerate(sorted(RETENTION.items())):
    r = 27 + k
    ws.cell(row=r, column=1, value=age).font = BODY
    c = ws.cell(row=r, column=2, value=ret)
    c.font, c.number_format = BLUE, '0.0%'
    for col in (1, 2):
        ws.cell(row=r, column=col).border = BORDER

ws["D26"] = "Other generating parameters"
ws["D26"].font = Font(name=ARIAL, bold=True, size=10)
others = [
    ("Mileage penalty", "exp(-0.045 x excess per 10k miles), clipped to [0.72, 1.22]"),
    ("Expected mileage for age", "10,000 miles per year"),
    ("Annual mileage drawn from", "Uniform(3,500, 21,000) - independent of age"),
    ("Price noise", "Normal(mean 1.0, sd 4.5%)"),
    ("Trim multiplier", "0.95 (Design) to 1.10 (Ultimate)"),
    ("Fuel multiplier", "0.88 (Electric) to 1.06 (Hybrid)"),
    ("Service history multiplier", "1.00 Full / 0.94 Partial / 0.87 None"),
    ("Seller multiplier", "1.08 Franchise / 1.00 Independent / 0.90 Private"),
    ("Condition multiplier", "1.04 Excellent / 1.00 Good / 0.92 Fair"),
    ("Former keeper penalty", "1.5% per keeper beyond the first"),
    ("Random seed", "20260826 - regenerating with this seed reproduces the file"),
]
for k, (a, b) in enumerate(others):
    r = 27 + k
    ws.cell(row=r, column=4, value=a).font = Font(name=ARIAL, size=10, bold=True)
    ws.cell(row=r, column=5, value=b).font = BODY

for col, w in zip("ABCDE", (34, 30, 34, 28, 58)):
    ws.column_dimensions[col].width = w

wb.save(OUT)

# Also emit a flat CSV with every derived column materialised. openpyxl writes
# formulas without cached values, so pandas reads those cells as blank; this is
# the file to load when fitting the model.
import csv
CSV_OUT = str(DATA_DIR / "vauxhall_astra_market_sample.csv")
with open(CSV_OUT, "w", newline="") as fh:
    w = csv.writer(fh)
    w.writerow(["listing_id", "reg_year", "age_years", "mileage",
                "avg_annual_mileage", "trim", "fuel", "transmission",
                "engine_litres", "body", "former_keepers", "service_history",
                "seller_type", "condition", "region", "original_list_gbp",
                "asking_price_gbp", "retained_pct"])
    for r in rows:
        (lid, reg, mil, trim, fuel, trans, eng, body, own, hist, sell, cond,
         region, listp, price) = r
        age = VALUATION_YEAR - reg
        w.writerow([lid, reg, age, mil,
                    "" if age == 0 else round(mil / age),
                    trim, fuel, trans, eng, body, own, hist, sell, cond,
                    region, listp, price, round(price / listp, 4)])

print("wrote", OUT)
print("wrote", CSV_OUT)
print("rows:", len(rows))
print("age range", int(ages.min()), "-", int(ages.max()))
print("price range", min(r[-1] for r in rows), "-", max(r[-1] for r in rows))
print("mileage range", min(r[2] for r in rows), "-", max(r[2] for r in rows))
corr = np.corrcoef([VALUATION_YEAR - r[1] for r in rows], [r[2] for r in rows])[0, 1]
print("corr(age, mileage) =", round(float(corr), 4))
